# -*- coding: utf-8 -*-
"""
Created on Tue Dec 19 13:07:37 2017

@author: chou
"""


from tkinter import messagebox
import openpyxl
import glob
import os


def Get_Batch_Info(self):
            
                #Get information from Physics files#
            
                if self.Option_Machine_var.get() == '':
                    self.Info_Window =  messagebox.showinfo(title = "Info", message = 'Please choose Machine.')
            
                if self.Option_Machine_var.get() == 'VG2':
            
                    os.chdir(r"P:\Production and Physics Communication\Tool Plan\VG")
                    
                    #Open file based on file name pattern#
                    for file_name in glob.glob('VG tool planning *.xlsm'):
                        self.Tool_Plan_name_VG = file_name
                
                    print(self.Tool_Plan_name_VG)
                    self.Plan_VG = openpyxl.load_workbook(self.Tool_Plan_name_VG)
                    self.Plan_VG_sheet = self.Plan_VG.get_sheet_by_name('Planning Tool')


                    #Specify the batch and get information#

                    for i in range(10000, 30000):
                        if self.Plan_VG_sheet.cell(row=i, column=2).value == self.var_Batch_Number.get() :
                            self.Batch_Location_Index = i
                            
                    
                    self.CW = self.Plan_VG_sheet.cell(row=self.Batch_Location_Index-1, column=2).value
                    self.Batch = self.Plan_VG_sheet.cell(row=self.Batch_Location_Index, column=2).value
                    self.Project = self.Plan_VG_sheet.cell(row=self.Batch_Location_Index+1, column=2).value
                    self.Aim = self.Plan_VG_sheet.cell(row=self.Batch_Location_Index-2, column=4).value          
                    self.Substrate_Number = self.Plan_VG_sheet.cell(row=self.Batch_Location_Index-1, column=10).value 
            
                    self.Substrate_list = []
                    for s in range (0, self.Substrate_Number):
                        self.Substrate_list.append(self.Plan_VG_sheet.cell(row=self.Batch_Location_Index+1+2*s, column=9).value)
            
            
                    self.Layer_count = 0
                    while self.Plan_VG_sheet.cell(row=self.Batch_Location_Index-1+self.Layer_count, column=4).value != None:
                        self.Layer_count = self.Layer_count + 1
                    
        
                    self.Architecture = []
                    for A in range (0,self.Layer_count):
                        for l in range (0,6):
                            self.Architecture.append(self.Plan_VG_sheet.cell(row=self.Batch_Location_Index-1+A, column=3+l).value)
                
                
                
                
                if self.Option_Machine_var.get() == 'Lesker':
            

                    os.chdir(r"P:\Production and Physics Communication\Tool Plan\Lesker")

                    for file_name in glob.glob('Lesker tool planning *.xlsm'):
                        self.Tool_Plan_name_Lesker = file_name
                        

                    print(self.Tool_Plan_name_Lesker)
                    self.Plan_Lekser = openpyxl.load_workbook(self.Tool_Plan_name_Lesker)

    
                    self.Plan_Lekser_sheet = self.Plan_Lekser.get_sheet_by_name('Planning Tool')


                    #Specify the batch#

                    for w in range(10000, 20000):
                        if self.Plan_Lekser_sheet.cell(row=w, column=2).value == self.var_Batch_Number.get() :
                            self.Batch_Location_Index = w
                    
                    self.CW = self.Plan_Lekser_sheet.cell(row=self.Batch_Location_Index-1, column=2).value
                    self.Batch = self.Plan_Lekser_sheet.cell(row=self.Batch_Location_Index, column=2).value
                    self.Project = self.Plan_Lekser_sheet.cell(row=self.Batch_Location_Index+1, column=2).value
                    self.Aim = self.Plan_Lekser_sheet.cell(row=self.Batch_Location_Index-2, column=4).value          
                    self.Substrate_Number = self.Plan_Lekser_sheet.cell(row=self.Batch_Location_Index-1, column=10).value 
            
                    self.Substrate_list = []
                    for s in range (0, self.Substrate_Number):
                        self.Substrate_list.append(self.Plan_Lekser_sheet.cell(row=self.Batch_Location_Index+1+2*s, column=9).value)
            
            
                    self.Layer_count = 0
                    while self.Plan_Lekser_sheet.cell(row=self.Batch_Location_Index-1+self.Layer_count, column=4).value != None:
                        self.Layer_count = self.Layer_count + 1
                    
        
                    self.Architecture = []
                    for A in range (0,self.Layer_count):
                        for l in range (0,6):
                            self.Architecture.append(self.Plan_Lekser_sheet.cell(row=self.Batch_Location_Index-1+A, column=3+l).value)
                
            
                
                self.Batch_Info= dict({'Calendar_Week' : self.CW, 'Batch_Number' : self.Batch, 'Project_Name' : self.Project, 'Aim' : self.Aim, 'Substrate_Number': self.Substrate_Number, 'Substrate_List' : self.Substrate_list, 'Layer_Number' : self.Layer_count, 'Architecture' : self.Architecture})
                
                return self.Batch_Info