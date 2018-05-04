# -*- coding: utf-8 -*-
"""
Created on Tue Dec 19 13:10:28 2017

@author: chou
"""

import os
import openpyxl

def Tooling_List_Info(self):
                
                if self.Option_Machine_var.get() == 'VG2':
                    
                    #Get information from Tooling list#
                    os.chdir(r"P:\Production\Tool_VG\Tooling VG & Status")
                
                    self.Tooling_list_name_VG = "Tooling list & status NEW VERSION.xlsx"  
                    self.Tooling_list_VG = openpyxl.load_workbook(self.Tooling_list_name_VG)
                    self.Tooling_list_VG_sheet = self.Tooling_list_VG.get_sheet_by_name('Tooling List ')
                    
                    
                    #Organic Chamber#
                    self.VG_Material_Name = []
                    self.VG_Material_ExpID = []
                    self.VG_Material_Toolings = []
                    self.VG_Material_Source = []
                    self.VG_Material_Sensor = []

                    #Organic Material#
                    for N in range(3,11):
                        self.VG_Material_Name.append(self.Tooling_list_VG_sheet.cell(row=N, column=1).value)
                        self.VG_Material_ExpID.append(self.Tooling_list_VG_sheet.cell(row=N, column=3).value)
                        self.VG_Material_Toolings.append(self.Tooling_list_VG_sheet.cell(row=N, column=9).value)
                        self.VG_Material_Source.append(self.Tooling_list_VG_sheet.cell(row=N, column=5).value)
                        self.VG_Material_Sensor.append(self.Tooling_list_VG_sheet.cell(row=N, column=6).value)


                    for N in range(12,22):
                        self.VG_Material_Name.append(self.Tooling_list_VG_sheet.cell(row=N, column=1).value)
                        self.VG_Material_ExpID.append(self.Tooling_list_VG_sheet.cell(row=N, column=3).value)
                        self.VG_Material_Toolings.append(self.Tooling_list_VG_sheet.cell(row=N, column=9).value)
                        self.VG_Material_Source.append(self.Tooling_list_VG_sheet.cell(row=N, column=5).value)
                        self.VG_Material_Sensor.append(self.Tooling_list_VG_sheet.cell(row=N, column=6).value)
                    


                    #Metal#
                    for N in range(28,32):
                        self.VG_Material_Name.append(self.Tooling_list_VG_sheet.cell(row=N, column=1).value)
                        self.VG_Material_ExpID.append(self.Tooling_list_VG_sheet.cell(row=N, column=3).value)
                        self.VG_Material_Toolings.append(self.Tooling_list_VG_sheet.cell(row=N, column=9).value)
                        self.VG_Material_Source.append(self.Tooling_list_VG_sheet.cell(row=N, column=5).value)
                        self.VG_Material_Sensor.append(self.Tooling_list_VG_sheet.cell(row=N, column=6).value)


                    
                    self.Tooling_List_Info = dict({'VG_Material_Name':self.VG_Material_Name,'VG_Material_ExpID':self.VG_Material_ExpID, 'VG_Material_Toolings':self.VG_Material_Toolings,'VG_Material_Source':self.VG_Material_Source,'VG_Material_Sensor':self.VG_Material_Sensor})
            
                    
            
                if self.Option_Machine_var.get() == 'Lesker':
            
                    #Get information from tooling list#
                
                    os.chdir(r"P:\Production\Tool_LESKER")
                
                    self.Tooling_list_name_Lesker = "Lesker tooling & status.xlsx"
                    self.Tooling_list_Lesker = openpyxl.load_workbook(self.Tooling_list_name_Lesker)
                    self.Tooling_list_Lesker_sheet = self.Tooling_list_Lesker.get_sheet_by_name('Chambers Status')
                    
                    self.Lesker_Material_Name = []
                    self.Lesker_Material_ExpID = []
                    self.Lesker_Material_Toolings = []
                    self.Lesker_Material_Source = []
                    self.Lesker_Material_Sensor = [1,1,2,2,3,3,4,4,5,5,6,6,7]
                    
                    #Organic Material#
                    for N in range(2,12):
                        self.Lesker_Material_Name.append(self.Tooling_list_Lesker_sheet.cell(row=N, column=4).value)
                        self.Lesker_Material_ExpID.append(self.Tooling_list_Lesker_sheet.cell(row=N, column=6).value)
                        self.Lesker_Material_Toolings.append(self.Tooling_list_Lesker_sheet.cell(row=N, column=3).value)
                        self.Lesker_Material_Source.append(self.Tooling_list_Lesker_sheet.cell(row=N, column=2).value)

                    #Metal Source#
                    for N in range(14,17):
                        self.Lesker_Material_Name.append(self.Tooling_list_Lesker_sheet.cell(row=N, column=4).value)
                        self.Lesker_Material_ExpID.append(self.Tooling_list_Lesker_sheet.cell(row=N, column=6).value)
                        self.Lesker_Material_Toolings.append(self.Tooling_list_Lesker_sheet.cell(row=N, column=3).value)
                        self.Lesker_Material_Source.append(self.Tooling_list_Lesker_sheet.cell(row=N, column=2).value)

                    


                    
                    self.Tooling_List_Info = dict({'Lesker_Material_Name':self.Lesker_Material_Name, 'Lesker_Material_ExpID':self.Lesker_Material_ExpID, 'Lesker_Material_Toolings':self.Lesker_Material_Toolings, 'Lesker_Material_Source':self.Lesker_Material_Source, 'Lesker_Material_Sensor':self.Lesker_Material_Sensor})
                    
                return self.Tooling_List_Info